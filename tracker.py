from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "DSA Tracker"

# 1. Main Static Left Headers (Row 4)
ws["A4"] = "Name"
ws["B4"] = "Solved"

# 2. Setup Problem Metadata Labels in Column B (Rows 1-3)
# This places the labels directly to the left of where the problems start
ws["B1"] = "Difficulty"
ws["B2"] = "Tags"
ws["B3"] = "Platform"

# 3. Add Problems Starting horizontally from Column C
problems = [
    {"diff": "EASY", "tag": "Arrays", "plat": "Leetcode", "title": "Sort Colors", "url": "https://leetcode.com"},
    {"diff": "Medium", "tag": "Math", "plat": "Leetcode", "title": "Largest Number", "url": "https://leetcode.com"}
]

# Loop to dynamically generate problem columns starting at Column 3 (C)
for i, prob in enumerate(problems):
    col_idx = 3 + i  # Column 3 is C, 4 is D, etc.
    
    # Write metadata vertically down the problem column
    ws.cell(row=1, column=col_idx, value=prob["diff"])
    ws.cell(row=2, column=col_idx, value=prob["tag"])
    ws.cell(row=3, column=col_idx, value=prob["plat"])
    
    # Write the Hyperlinked Problem Title in Row 4
    cell = ws.cell(row=4, column=col_idx, value=prob["title"])
    cell.hyperlink = prob["url"]
    cell.font = Font(color="0000FF", underline="single") # Standard Blue link

# 4. Populate Member Names starting at Row 12
members = ["Mekhluqat Abdulwehab", "Nezira Werku", "sunan", "mscoder"]

for row_idx, name in enumerate(members, start=12):
    ws.cell(row=row_idx, column=1, value=name) # Column A = Name
    ws.cell(row=row_idx, column=2, value=0)    # Column B = Solved Counter (Defaults to 0)
    
    # Initialize data tracking tracking cells for individual problems (Columns C onwards)
    for col_idx in range(3, 3 + len(problems)):
        ws.cell(row=row_idx, column=col_idx, value=0)

wb.save("dsa_tracker.xlsx")
